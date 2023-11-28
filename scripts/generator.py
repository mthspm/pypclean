import hashlib
import random
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom
import json
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from utils import *

class LogGenerator:
    """
    LogGenerator
    Class responsible for generating logs in different formats.
    @param: None
    @return: None
    """
    def __init__(self):
        pass

    def hash_generator(self, seed, size=16):
        random.seed(seed)
        return hashlib.sha256(str(random.getrandbits(256)).encode('utf-8')).hexdigest()[:size]
    
    def run(self, data, format, destination):
        if format.lower() == 'pdf':
            self.generate_pdf(data, destination)
        elif format.lower() == 'txt':
            self.generate_txt(data, destination)
        elif format.lower() == 'excel':
            self.generate_excel(data, destination)
        elif format.lower() == 'xml':
            self.generate_xml(data, destination)
        elif format.lower() == 'json':
            self.generate_json(data, destination)
        else:
            return False

    def generate_pdf(self, data, destination):

        # Setup
        x = 30
        xi = 420
        y = 750
        seed = self.hash_generator(time.time())
        random_hash = self.hash_generator(seed)
        c = canvas.Canvas(f"{destination}/pcscan_log_{random_hash}.pdf")
        c.setPageSize((letter[0], letter[1]))
        c.setFont("Helvetica", 12)

        # Header
        c.drawString(x, y, "PC SCAN 0.1")
        c.drawString(x, y - 15, "Relatório de Informações do Sistema, Pagina 1")
        c.drawString(xi, y, f"Data: {timenow()}")
        c.drawString(xi, y - 15, f"Hash: {random_hash}")

        # Builders
        def add_section(section_title, data):
            nonlocal temp_y
            c.drawString(x, temp_y, section_title)
            c.drawString(x, temp_y - 15, "----------------------------------------")
            temp_y -= 30
            for key, value in data.items():
                c.drawString(x, temp_y, f"{key} : {value}")
                temp_y -= 15
            c.drawString(x, temp_y, "----------------------------------------")
            temp_y -= 15

        def add_section_right(section_title, data):
            nonlocal temp_yi
            c.drawString(420, temp_yi, section_title)
            c.drawString(420, temp_yi - 15, "----------------------------------------")
            temp_yi -= 30
            for key, value in data.items():
                c.drawString(420, temp_yi, f"{key} : {value}")
                temp_yi -= 15
            c.drawString(420, temp_yi, "----------------------------------------")
            temp_yi -= 15

        temp_y = y - 45
        temp_yi = y - 45
        add_section("Sistema", data['system'])
        add_section("CPU", data['cpu'])
        add_section("Memória", data['memory'])
        add_section("GPU", data['gpu'])
        add_section("Disco", data['disk'])
        add_section_right("Rede", data['network'])
        add_section_right("Placa Mãe", data['motherboard'])

        c.drawString(xi, temp_yi, "END OF FILE - PC SCAN 0.1")
        c.drawString(xi, temp_yi - 15, "DRIVERS AVAIABLE IN PG2")

        # Drivers
        c.showPage()
        c.drawString(x, y, "PC SCAN 0.1")
        c.drawString(x, y - 15, "Relatório de Drivers do Sistema, Pagina 2")
        c.drawString(xi, y, f"Data: {timenow()}")
        c.drawString(xi, y - 15, f"Hash: {random_hash}")

        c.save()

    def generate_txt(self, data, destination):
        seed = self.hash_generator(time.time())
        random_hash = self.hash_generator(seed)
        with open(f"{destination}/pcscan_log_{random_hash}.txt", "w") as f:
            f.write(f"PC SCAN 0.1\n")
            f.write(f"Relatório de Informações do Sistema, Pagina 1\n")
            f.write(f"Data: {timenow()}\n")
            f.write(f"Hash: {random_hash}\n\n")

            f.write(f"Sistema\n")
            f.write(f"----------------------------------------\n")
            for key, value in data['system'].items():
                f.write(f"{key} : {value}\n")
            f.write(f"----------------------------------------\n\n")

            f.write(f"CPU\n")
            f.write(f"----------------------------------------\n")
            for key, value in data['cpu'].items():
                f.write(f"{key} : {value}\n")
            f.write(f"----------------------------------------\n\n")

            f.write(f"Memória\n")
            f.write(f"----------------------------------------\n")
            for key, value in data['memory'].items():
                f.write(f"{key} : {value}\n")
            f.write(f"----------------------------------------\n\n")

            f.write(f"GPU\n")
            f.write(f"----------------------------------------\n")
            for key, value in data['gpu'].items():
                f.write(f"{key} : {value}\n")
            f.write(f"----------------------------------------\n\n")
            f.write(f"Disco\n")
            f.write(f"----------------------------------------\n")
            for key, value in data['disk'].items():
                f.write(f"{key} : {value}\n")
            f.write(f"----------------------------------------\n\n")
            f.write(f"Rede\n")
            f.write(f"----------------------------------------\n")
            for key, value in data['network'].items():
                f.write(f"{key} : {value}\n")
            f.write(f"----------------------------------------\n\n")
            f.write(f"Placa Mãe\n")
            f.write(f"----------------------------------------\n")
            for key, value in data['motherboard'].items():
                f.write(f"{key} : {value}\n")
            f.write(f"----------------------------------------\n\n")
            f.write(f"Drivers\n")
            f.write(f"----------------------------------------\n")
            for key in data['drivers']:
                f.write(f"{key}\n")
                f.write(f"----------------------------------------\n")
            f.write(f"END OF FILE - PC SCAN 0.1\n")
            f.write("THANK YOU FOR USING OUR SOFTWARE!\n")
            f.close()

    def generate_xml(self, data, destination):
        seed = self.hash_generator(time.time())
        random_hash = self.hash_generator(seed)
        
        def generate_xml_element(root, category, data):
            element = ET.SubElement(root, category)
            element.text = str(data)

        root = ET.Element("pcscan_log")

        generate_xml_element(root, "hash", random_hash)
        generate_xml_element(root, "date", timenow())

        categories = ["system", "cpu", "memory", "gpu", "disk", "network", "motherboard", "drivers"]
        for category in categories:
            generate_xml_element(root, category, data[category])

        xml_string = minidom.parseString(ET.tostring(root)).toprettyxml(indent="    ")

        with open(f"{destination}/pcscan_log_{random_hash}.xml", "w") as arquivo_xml:
            arquivo_xml.write(xml_string)

    def generate_json(self, data, destination):
        seed = self.hash_generator(time.time())
        random_hash = self.hash_generator(seed)
        with open(f"{destination}/pcscan_log_{random_hash}.json", "w") as f:
            json.dump(data, f, indent=4)
            f.close()

    def generate_excel(self, data, destination):
        # Setup
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        def set_font(cell, bold=False):
            cell.font = openpyxl.styles.Font(bold=bold)

        # Header Setup
        header = [sheet["A1"], sheet["A2"], sheet["B1"], sheet["B2"], sheet["C1"], sheet["D1"], sheet["C2"], sheet["D2"]]
        sheet["A1"] = "PCSCAN"
        sheet["A2"] = "V0.1"
        sheet["B1"] = "Relatório de Informações do Sistema"
        sheet["B2"] = "github.com/mathlokz/pcscan"
        sheet["C1"] = "Data:"
        sheet["D1"] = timenow()
        hash_value = self.hash_generator(time.time())
        sheet["C2"] = "Hash:"
        sheet["D2"] = hash_value
        for cell in header:
            set_font(cell, bold=True)
        
        # Body Setup
        categories = ['system', 'cpu', 'memory', 'gpu', 'disk', 'network', 'motherboard']
        row = 6

        for category in categories:
            sheet[f"A{row}"] = category.upper()
            set_font(sheet[f"A{row}"], bold=True)
            row += 1
            for key, value in data[category].items():
                sheet[f"A{row}"] = f"{key}"
                sheet[f"B{row}"] = f"{value}"
                row += 1
            row += 1

        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # Converte o número da coluna em letra
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        file_name = f"{destination}/pcscan_log_{hash_value}.xlsx"
        workbook.save(file_name)